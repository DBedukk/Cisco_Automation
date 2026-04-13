
"""
Copyright (c) 2023 Cisco and/or its affiliates.

This software is licensed to you under the terms of the Cisco Sample
Code License, Version 1.1 (the "License"). You may obtain a copy of the
License at

               https://developer.cisco.com/docs/licenses

All use of the material herein must be in accordance with the terms of
the License. All rights not expressly granted by the License are
reserved. Unless required by applicable law or agreed to separately in
writing, software distributed under the License is distributed on an "AS
IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
or implied.

"""

__author__ = "Doruk Beduk"
__copyright__ = "Copyright (c) 2026 Cisco and/or its affiliates."
__license__ = "Cisco Sample Code License, Version 1.1"

import json
import os
import sys
import time
from argparse import ArgumentParser
from datetime import datetime
from getpass import getpass

import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, GradientFill, PatternFill, Side)
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Rich imports ──────────────────────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn, TimeElapsedColumn
    from rich.prompt import Confirm, Prompt
    from rich.table import Table
    from rich import box
    from rich.text import Text
    from rich.rule import Rule
    from rich.columns import Columns
    from rich.live import Live
    from rich.align import Align
except ImportError:
    print("ERROR: 'rich' library is required. Install it with:  pip install rich")
    sys.exit(1)

console = Console()

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')

# ── Object types to enumerate ─────────────────────────────────────────────────
# Each entry: (api_path_segment, friendly_label, csv_columns_fn)
# csv_columns_fn receives one item dict and returns an ordered dict of columns.

OBJECT_TYPES = [
    ("hosts",               "Host Objects"),
    ("networks",            "Network Objects"),
    ("ranges",              "Range Objects"),
    ("networkgroups",       "Network Groups"),
    ("fqdns",               "FQDN Objects"),
    ("urls",                "URL Objects"),
    ("urlgroups",           "URL Groups"),
    ("portobjectgroups",    "Port Object Groups"),
    ("protocolportobjects", "Protocol Port Objects"),
    ("icmpv4objects",       "ICMPv4 Objects"),
    ("icmpv6objects",       "ICMPv6 Objects"),
    ("securityzones",       "Security Zones"),
    ("interfacegroups",     "Interface Groups"),
    ("vlantags",            "VLAN Tags"),
    ("vlangrouptags",       "VLAN Group Tags"),
    ("applicationfilters",  "Application Filters"),
    ("timeranges",          "Time Ranges"),
    ("geolocation",         "Geolocation Objects"),
    ("variablesets",        "Variable Sets"),
    ("tunneltags",          "Tunnel Tags"),
    ("slamonitors",         "SLA Monitors"),
]


# ──────────────────────────────────────────────────────────────────────────────
#  Authentication helpers
# ──────────────────────────────────────────────────────────────────────────────

def authenticate(server: str, username: str, password: str, cert_path=None) -> dict:
    """POST to generatetoken; return a dict with token, refresh_token, domain_uuid."""
    url = f"https://{server}/api/fmc_platform/v1/auth/generatetoken"
    headers = {"Content-Type": "application/json"}
    verify = cert_path if cert_path else False

    try:
        resp = requests.post(
            url, headers=headers,
            auth=requests.auth.HTTPBasicAuth(username, password),
            verify=verify, timeout=30
        )
    except requests.exceptions.ConnectionError as exc:
        console.print(f"[bold red]Connection error:[/bold red] {exc}")
        sys.exit(1)
    except requests.exceptions.Timeout:
        console.print("[bold red]Connection timed out.[/bold red]")
        sys.exit(1)

    if resp.status_code not in (200, 204):
        console.print(f"[bold red]Authentication failed[/bold red] — HTTP {resp.status_code}")
        console.print(resp.text[:500])
        sys.exit(1)

    token         = resp.headers.get("X-auth-access-token", "")
    refresh_token = resp.headers.get("X-auth-refresh-token", "")
    domain_uuid   = resp.headers.get("DOMAIN_UUID", "")

    if not token:
        console.print("[bold red]No access token returned. Check credentials.[/bold red]")
        sys.exit(1)

    return {"token": token, "refresh_token": refresh_token, "domain_uuid": domain_uuid}


def logout(server: str, token: str, cert_path=None):
    url = f"https://{server}/api/fmc_platform/v1/auth/revokeaccess"
    headers = {"Content-Type": "application/json", "X-auth-access-token": token}
    verify = cert_path if cert_path else False
    try:
        requests.post(url, headers=headers, verify=verify, timeout=10)
    except Exception:
        pass


# ──────────────────────────────────────────────────────────────────────────────
#  Domain discovery
# ──────────────────────────────────────────────────────────────────────────────

def get_all_domains(server: str, token: str, cert_path=None) -> list[dict]:
    """Return a flat list of all domains (Global + sub-domains) from FMC."""
    url = f"https://{server}/api/fmc_platform/v1/info/domain"
    headers = {
        "Content-Type": "application/json",
        "X-auth-access-token": token,
    }
    verify = cert_path if cert_path else False
    domains = []
    offset, limit = 0, 1000

    while True:
        resp = requests.get(
            url, headers=headers,
            params={"limit": limit, "offset": offset, "expanded": True},
            verify=verify, timeout=30
        )
        if resp.status_code != 200:
            console.print(f"[bold red]Failed to retrieve domains:[/bold red] HTTP {resp.status_code}")
            console.print(resp.text[:500])
            sys.exit(1)

        data = resp.json()
        items = data.get("items", [])
        domains.extend(items)

        if len(items) < limit:
            break
        offset += limit

    return domains




# Object types that support filter=unusedOnly:true per the FMC REST API docs.
# Uses the combined endpoints where documented (networkaddresses covers hosts+networks,
# ports covers all protocol/port objects).
UNUSED_OBJECT_TYPES = [
    ("networkaddresses",    "Network & Host Objects"),   # hosts + networks combined
    ("ports",               "Port Objects"),              # all port-type objects combined
    ("portobjectgroups",    "Port Object Groups"),
    ("networkgroups",       "Network Groups"),
    ("ranges",              "Range Objects"),
    ("fqdns",               "FQDN Objects"),
    ("urls",                "URL Objects"),
    ("urlgroups",           "URL Groups"),
    ("icmpv4objects",       "ICMPv4 Objects"),
    ("icmpv6objects",       "ICMPv6 Objects"),
]


# ──────────────────────────────────────────────────────────────────────────────
#  Object fetching
# ──────────────────────────────────────────────────────────────────────────────

def fetch_objects(server: str, token: str, domain_uuid: str, obj_type: str,
                  cert_path=None) -> list[dict]:
    """Paginate through all items for a given object type in a domain."""
    url = f"https://{server}/api/fmc_config/v1/domain/{domain_uuid}/object/{obj_type}"
    headers = {
        "Content-Type": "application/json",
        "X-auth-access-token": token,
    }
    verify = cert_path if cert_path else False
    items = []
    offset, limit = 0, 1000

    while True:
        try:
            resp = requests.get(
                url, headers=headers,
                params={"limit": limit, "offset": offset, "expanded": True},
                verify=verify, timeout=30
            )
        except requests.exceptions.Timeout:
            break

        if resp.status_code == 401:
            # Token expired — caller should handle refresh; for now just stop.
            break
        if resp.status_code == 404:
            # Object type not supported in this domain/version — skip silently.
            break
        if resp.status_code == 405:
            break
        if resp.status_code != 200:
            break

        data = resp.json()
        batch = data.get("items", [])
        items.extend(batch)

        if len(batch) < limit:
            break
        offset += limit
        time.sleep(0.15)   # stay well under 120 req/min rate limit

    return items


def fetch_unused_objects(server: str, token: str, domain_uuid: str, obj_type: str,
                         cert_path=None) -> list[dict]:
    """
    Fetch objects not referenced by any policy in the queried domain.
    Uses filter=unusedOnly:true per the FMC REST API v7.0 docs.
    """
    url = f"https://{server}/api/fmc_config/v1/domain/{domain_uuid}/object/{obj_type}"
    headers = {
        "Content-Type": "application/json",
        "X-auth-access-token": token,
    }
    verify = cert_path if cert_path else False
    items = []
    offset, limit = 0, 1000

    while True:
        try:
            resp = requests.get(
                url, headers=headers,
                params={"limit": limit, "offset": offset, "expanded": True, "filter": "unusedOnly:true"},
                verify=verify, timeout=30
            )
        except requests.exceptions.Timeout:
            break

        if resp.status_code in (401, 404, 405):
            break
        if resp.status_code != 200:
            break

        data = resp.json()
        batch = data.get("items", [])
        items.extend(batch)

        if len(batch) < limit:
            break
        offset += limit
        time.sleep(0.15)

    return items


# ──────────────────────────────────────────────────────────────────────────────
#  Data model
# ──────────────────────────────────────────────────────────────────────────────

# Sheet definitions: (sheet_title, [object type labels that belong here], [column headers])
SHEET_DEFINITIONS = [
    (
        "Host Objects",
        {"Host Objects"},
        ["Domain", "Name", "IP Address", "Description", "Overridable", "Last Modified By", "Last Modified"],
    ),
    (
        "Network Objects",
        {"Network Objects", "Range Objects", "Network Groups", "FQDN Objects"},
        ["Domain", "Object Type", "Name", "Value / Members", "Description", "Overridable", "Last Modified By", "Last Modified"],
    ),
    (
        "Services",
        {"Protocol Port Objects", "Port Object Groups", "ICMPv4 Objects", "ICMPv6 Objects"},
        ["Domain", "Object Type", "Name", "Protocol", "Port / Code", "Members", "Description", "Last Modified By", "Last Modified"],
    ),
    (
        "URL Objects",
        {"URL Objects", "URL Groups"},
        ["Domain", "Object Type", "Name", "URL / Members", "Description", "Overridable", "Last Modified By", "Last Modified"],
    ),
    (
        "Other Objects",
        None,   # catches everything not in the sheets above
        ["Domain", "Object Type", "Name", "Value", "Members", "Description", "Overridable", "Last Modified By", "Last Modified"],
    ),
    (
        "Unused Objects",
        None,   # populated separately via unusedonly=true API calls
        ["Domain", "Object Type", "Name", "Value / Members", "Description", "Overridable", "Last Modified By", "Last Modified"],
    ),
]

# All label sets that have dedicated sheets (used to route "Other Objects")
_DEDICATED_LABELS: set[str] = set()
for _, labels, _ in SHEET_DEFINITIONS[:-2]:   # exclude Other + Unused (both None)
    if labels:
        _DEDICATED_LABELS.update(labels)


# ── Colours ───────────────────────────────────────────────────────────────────
_CISCO_BLUE   = "00538B"   # Cisco brand blue
_HEADER_TEXT  = "FFFFFF"
_ROW_ALT      = "EEF4FA"   # light blue-grey stripe
_ROW_WHITE    = "FFFFFF"
_ACCENT_GREEN = "00A86B"   # summary header accent
_GRID_COLOUR  = "C5D7E8"


def _border(style="thin") -> Border:
    s = Side(style=style, color=_GRID_COLOUR)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _parse_timestamp(ts) -> str:
    """Convert FMC millisecond epoch timestamp to YYYY-MM-DD HH:MM:SS."""
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts) / 1000).strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, OSError, OverflowError):
        return str(ts)


def _flatten_value(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (list, dict)):
        return json.dumps(val, separators=(",", ":"))
    return str(val)


# ── Row builder ───────────────────────────────────────────────────────────────

def _row_for_item(item: dict, obj_type_label: str, domain_name: str) -> dict:
    """Build a normalised data dict for one FMC object."""
    meta = item.get("metadata", {}) or {}

    value = (
        item.get("value")
        or item.get("prefix")
        or item.get("url")
        or ""
    )

    if "objects" in item:
        members = "; ".join(o.get("name", o.get("id", "")) for o in item["objects"])
    elif "literals" in item:
        members = "; ".join(_flatten_value(l) for l in item["literals"])
    else:
        members = ""

    last_user = ""
    lu = meta.get("lastUser")
    if isinstance(lu, dict):
        last_user = lu.get("name", "")
    elif isinstance(lu, str):
        last_user = lu

    return {
        "Domain":       domain_name,
        "ObjectType":   obj_type_label,
        "Name":         item.get("name", ""),
        "ID":           item.get("id", ""),
        "Description":  item.get("description", "") or "",
        "Value":        str(value),
        "Members":      members,
        "Overridable":  "Yes" if item.get("overridable") else "No",
        "Protocol":     item.get("protocol", "") or "",
        "Port":         str(item.get("port", "") or ""),
        "InterfaceMode": item.get("interfaceMode", "") or "",
        "LastUser":     last_user,
        "LastModified": str(meta.get("timestamp", "") or ""),
    }


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _style_header_row(ws, row_num: int, num_cols: int, fill_colour: str):
    """Apply bold white header styling to a worksheet row."""
    fill   = _header_fill(fill_colour)
    font   = Font(bold=True, color=_HEADER_TEXT, name="Calibri", size=11)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = _border("medium")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border


def _style_data_row(ws, row_num: int, num_cols: int, alternate: bool):
    fill   = _header_fill(_ROW_ALT if alternate else _ROW_WHITE)
    font   = Font(name="Calibri", size=10)
    align  = Alignment(vertical="center", wrap_text=False)
    border = _border("thin")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
        cell.border    = border


def _auto_fit_columns(ws, min_width=12, max_width=55):
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 3, max_width))


def _write_data_sheet(ws, headers: list[str], rows: list[list], fill_colour: str):
    """Write headers + data rows with full styling."""
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Header
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(headers), fill_colour)

    # Data
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        _style_data_row(ws, row_idx, len(headers), alternate=(row_idx % 2 == 0))

    _auto_fit_columns(ws)


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _write_summary_sheet(ws, summary: dict, generated_at: str, domains_queried: list[str]):
    ws.freeze_panes = "A2"

    # Title block
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = "FMC Object Export — Summary"
    title_cell.font      = Font(bold=True, size=14, color=_HEADER_TEXT, name="Calibri")
    title_cell.fill      = _header_fill(_CISCO_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Meta rows
    meta_rows = [
        ("Generated",        generated_at),
        ("Domains Queried",  ", ".join(domains_queried)),
    ]
    for i, (label, val) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=i, column=2, value=val).font   = Font(name="Calibri", size=10)

    # Spacer
    spacer_row = 2 + len(meta_rows) + 1

    # Sub-header
    headers = ["Domain", "Object Type", "Count"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=spacer_row, column=col_idx, value=h)
    _style_header_row(ws, spacer_row, len(headers), _ACCENT_GREEN)
    ws.row_dimensions[spacer_row].height = 20

    # Data
    data_row = spacer_row + 1
    grand_total = 0
    for domain_name, type_counts in summary.items():
        for obj_type, count in sorted(type_counts.items()):
            if count == 0:
                continue
            ws.cell(row=data_row, column=1, value=domain_name)
            ws.cell(row=data_row, column=2, value=obj_type)
            ws.cell(row=data_row, column=3, value=count)
            _style_data_row(ws, data_row, 3, alternate=(data_row % 2 == 0))
            grand_total += count
            data_row += 1

    # Grand total
    for col, val in [(1, "TOTAL"), (2, ""), (3, grand_total)]:
        cell = ws.cell(row=data_row, column=col, value=val)
        cell.font   = Font(bold=True, name="Calibri", size=10, color=_HEADER_TEXT)
        cell.fill   = _header_fill(_CISCO_BLUE)
        cell.border = _border("medium")
        cell.alignment = Alignment(horizontal="center" if col == 3 else "left", vertical="center")

    _auto_fit_columns(ws)


# ── Main export entry point ───────────────────────────────────────────────────

def write_excel(all_rows: list[dict], unused_rows: list[dict], summary: dict,
                filepath: str, domains_queried: list[str], generated_at: str):
    """Write a professional multi-sheet Excel workbook."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    wb = Workbook()

    # ── Summary sheet (first tab) ─────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, summary, generated_at, domains_queried)

    # ── Route rows to sheets ──────────────────────────────────────────────────
    sheet_rows: dict[str, list] = {sd[0]: [] for sd in SHEET_DEFINITIONS}

    for row in all_rows:
        label = row["ObjectType"]
        placed = False
        for sheet_title, label_set, _ in SHEET_DEFINITIONS[:-1]:
            if label_set and label in label_set:
                sheet_rows[sheet_title].append(row)
                placed = True
                break
        if not placed:
            sheet_rows["Other Objects"].append(row)

    # ── Sheet colours ─────────────────────────────────────────────────────────
    sheet_colours = {
        "Host Objects":    _CISCO_BLUE,
        "Network Objects": "1A6E37",   # dark green
        "Services":        "8B3A00",   # dark orange-brown
        "URL Objects":     "6B008B",   # purple
        "Other Objects":   "3D3D3D",   # dark grey
        "Unused Objects":  "8B0000",   # dark red — unused = potential cleanup
    }

    # ── Column value extractors per sheet ─────────────────────────────────────
    def host_cols(r):
        return [r["Domain"], r["Name"], r["Value"], r["Description"],
                r["Overridable"], r["LastUser"], r["LastModified"]]

    def network_cols(r):
        val = r["Value"] if r["Value"] else r["Members"]
        return [r["Domain"], r["ObjectType"], r["Name"], val,
                r["Description"], r["Overridable"], r["LastUser"], r["LastModified"]]

    def service_cols(r):
        return [r["Domain"], r["ObjectType"], r["Name"], r["Protocol"],
                r["Port"], r["Members"], r["Description"], r["LastUser"], r["LastModified"]]

    def url_cols(r):
        val = r["Value"] if r["Value"] else r["Members"]
        return [r["Domain"], r["ObjectType"], r["Name"], val,
                r["Description"], r["Overridable"], r["LastUser"], r["LastModified"]]

    def other_cols(r):
        return [r["Domain"], r["ObjectType"], r["Name"], r["Value"],
                r["Members"], r["Description"], r["Overridable"], r["LastUser"], r["LastModified"]]

    extractors = {
        "Host Objects":    host_cols,
        "Network Objects": network_cols,
        "Services":        service_cols,
        "URL Objects":     url_cols,
        "Other Objects":   other_cols,
        "Unused Objects":  lambda r: [
            r["Domain"],
            r["ObjectType"],
            r["Name"],
            r["Value"] if r["Value"] else r["Members"],
            r["Description"],
            r["Overridable"],
            r["LastUser"],
            r["LastModified"],
        ],
    }

    # ── Write each sheet (skip Unused Objects — handled separately below) ─────
    for sheet_title, _, headers in SHEET_DEFINITIONS[:-1]:
        rows_for_sheet = sheet_rows[sheet_title]
        if not rows_for_sheet:
            continue   # skip empty sheets

        ws = wb.create_sheet(title=sheet_title)
        colour = sheet_colours[sheet_title]
        extractor = extractors[sheet_title]
        data = [extractor(r) for r in rows_for_sheet]
        _write_data_sheet(ws, headers, data, colour)

    # ── Unused Objects sheet (always created) ────────────────────────────────
    _, _, unused_headers = SHEET_DEFINITIONS[-1]
    ws_unused = wb.create_sheet(title="Unused Objects")
    unused_extractor = extractors["Unused Objects"]

    if unused_rows:
        unused_data = [unused_extractor(r) for r in unused_rows]
        _write_data_sheet(ws_unused, unused_headers, unused_data, sheet_colours["Unused Objects"])
    else:
        # Write the header row then a single informational row
        _write_data_sheet(ws_unused, unused_headers, [], sheet_colours["Unused Objects"])
        no_data_cell = ws_unused.cell(row=2, column=1,
                                      value="No unused objects found in the queried domain(s).")
        no_data_cell.font      = Font(italic=True, color="666666", name="Calibri", size=10)
        no_data_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_unused.merge_cells(start_row=2, start_column=1,
                              end_row=2, end_column=len(unused_headers))

    wb.save(filepath)


# ──────────────────────────────────────────────────────────────────────────────
#  UI helpers
# ──────────────────────────────────────────────────────────────────────────────

def print_banner():
    banner = Panel(
        Align(
            Text("FMC Object Parser", style="bold cyan", justify="center"),
            align="center"
        ),
        subtitle="[dim]Cisco Firepower Management Center — Domain Object Exporter[/dim]",
        border_style="cyan",
        padding=(1, 4),
    )
    console.print(banner)
    console.print()


def build_domain_tree(domains: list[dict]) -> list[dict]:
    """
    Sort domains so Global appears first, then sub-domains alphabetically.
    Attach a display_name with indentation to reflect hierarchy.
    """
    # FMC returns a 'name' like 'Global' or 'Global/Child' or 'Global/Child/GrandChild'
    def depth(d):
        return d.get("name", "").count("/")

    sorted_domains = sorted(domains, key=lambda d: (depth(d), d.get("name", "").lower()))

    for d in sorted_domains:
        parts = d.get("name", "").split("/")
        indent = "  " * (len(parts) - 1)
        leaf   = parts[-1]
        d["display_name"] = f"{indent}{leaf}"

    return sorted_domains


def display_domain_table(domains: list[dict]) -> Table:
    table = Table(
        title="[bold cyan]Available Domains[/bold cyan]",
        box=box.ROUNDED,
        header_style="bold white on #1a1a4e",
        border_style="cyan",
        show_lines=True,
        expand=False,
    )
    table.add_column("#",           style="bold yellow",  justify="right",  no_wrap=True)
    table.add_column("Domain Name", style="bold white",   justify="left",   min_width=30)
    table.add_column("UUID",        style="dim",          justify="left",   min_width=36)
    table.add_column("Type",        style="green",        justify="center", no_wrap=True)

    for idx, d in enumerate(domains, start=1):
        parts = d.get("name", "").split("/")
        is_global = len(parts) == 1
        dtype = "[bold magenta]Global[/bold magenta]" if is_global else "[cyan]Sub-Domain[/cyan]"
        table.add_row(str(idx), d["display_name"], d.get("uuid", ""), dtype)

    return table


def prompt_domain_selection(domains: list[dict]) -> list[dict]:
    """
    Display the table, then ask the user which domain(s) to parse.
    Supports: a single number, a comma-separated list, a range (e.g. 2-5), or 'all'.
    """
    table = display_domain_table(domains)
    console.print(table)
    console.print()
    console.print("[dim]Enter domain number(s) to parse. Examples:[/dim]")
    console.print("  [yellow]1[/yellow]        — single domain")
    console.print("  [yellow]1,3,5[/yellow]    — multiple domains")
    console.print("  [yellow]2-6[/yellow]      — range of domains")
    console.print("  [yellow]all[/yellow]      — every domain listed above")
    console.print()

    while True:
        raw = Prompt.ask("[bold cyan]Your selection[/bold cyan]").strip().lower()

        if raw == "all":
            return domains

        selected_indices = set()
        try:
            for part in raw.split(","):
                part = part.strip()
                if "-" in part:
                    lo, hi = part.split("-", 1)
                    for i in range(int(lo), int(hi) + 1):
                        selected_indices.add(i)
                else:
                    selected_indices.add(int(part))
        except ValueError:
            console.print("[red]Invalid input — please try again.[/red]")
            continue

        bad = [i for i in selected_indices if i < 1 or i > len(domains)]
        if bad:
            console.print(f"[red]Out-of-range numbers: {bad}. Max is {len(domains)}.[/red]")
            continue

        chosen = [domains[i - 1] for i in sorted(selected_indices)]
        return chosen


def display_summary_table(results: dict):
    """results: {domain_name: {obj_type: count}}"""
    table = Table(
        title="[bold green]Export Summary[/bold green]",
        box=box.ROUNDED,
        header_style="bold white on dark_green",
        border_style="green",
        show_lines=True,
    )
    table.add_column("Domain",      style="bold white", min_width=25)
    table.add_column("Object Type", style="cyan",       min_width=22)
    table.add_column("Count",       style="yellow",     justify="right")

    grand_total = 0
    for domain_name, type_counts in results.items():
        first = True
        for obj_type, count in sorted(type_counts.items()):
            if count == 0:
                continue
            d_col = f"[bold]{domain_name}[/bold]" if first else ""
            table.add_row(d_col, obj_type, str(count))
            grand_total += count
            first = False

    table.add_section()
    table.add_row("[bold]TOTAL[/bold]", "", f"[bold yellow]{grand_total}[/bold yellow]")
    console.print(table)


# ──────────────────────────────────────────────────────────────────────────────
#  Core orchestration
# ──────────────────────────────────────────────────────────────────────────────

def parse_domains(server: str, token: str, chosen_domains: list[dict],
                  cert_path=None) -> tuple[list[dict], list[dict], dict]:
    """
    Fetch all objects and unused objects across chosen_domains.
    Returns (all_rows, unused_rows, summary_dict).
    """
    all_rows    = []
    unused_rows = []
    summary     = {}  # {domain_name: {obj_type_label: count}}

    with Progress(
        SpinnerColumn(),
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(bar_width=40),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=False,
    ) as progress:

        domain_task = progress.add_task("[cyan]Domains[/cyan]", total=len(chosen_domains))

        for domain in chosen_domains:
            domain_name = domain.get("name", "Unknown")
            domain_uuid = domain.get("uuid", "")
            summary[domain_name] = {}

            obj_task = progress.add_task(
                f"  [white]{domain_name}[/white]",
                total=len(OBJECT_TYPES) + len(UNUSED_OBJECT_TYPES)
            )

            # ── All objects ───────────────────────────────────────────────────
            for obj_type, label in OBJECT_TYPES:
                progress.update(obj_task, description=f"  [white]{domain_name}[/white] → [cyan]{label}[/cyan]")
                items = fetch_objects(server, token, domain_uuid, obj_type, cert_path)

                rows_for_type = [_row_for_item(item, label, domain_name) for item in items]
                all_rows.extend(rows_for_type)
                summary[domain_name][label] = len(rows_for_type)
                progress.advance(obj_task)

            # ── Unused objects ────────────────────────────────────────────────
            for obj_type, label in UNUSED_OBJECT_TYPES:
                progress.update(obj_task, description=f"  [white]{domain_name}[/white] → [red]Unused {label}[/red]")
                items = fetch_unused_objects(server, token, domain_uuid, obj_type, cert_path)

                rows_for_type = [_row_for_item(item, label, domain_name) for item in items]
                unused_rows.extend(rows_for_type)
                progress.advance(obj_task)

            progress.update(obj_task, description=f"  [green]✓ {domain_name}[/green]")
            progress.advance(domain_task)

    return all_rows, unused_rows, summary


# ──────────────────────────────────────────────────────────────────────────────
#  Entry point
# ──────────────────────────────────────────────────────────────────────────────

def build_args():
    parser = ArgumentParser(
        description="FMC Object Parser — export FMC objects to CSV by domain",
        formatter_class=lambda prog: __import__("argparse").HelpFormatter(prog, max_help_position=40),
    )
    parser.add_argument("-s", "--server",    required=True,  help="FMC IP or hostname")
    parser.add_argument("-u", "--username",  required=True,  help="FMC API username")
    parser.add_argument("-p", "--password",  default=None,   help="FMC password (prompted if omitted)")
    parser.add_argument("-c", "--cert-path", default=None,   help="Path to CA bundle (omit to skip TLS verification)")
    parser.add_argument("-o", "--output",    default=None,  help="Output CSV filename (default: auto-generated)")
    parser.add_argument("--no-logout",       action="store_true", help="Skip revoking the API token on exit")
    return parser.parse_args()


def main():
    print_banner()
    args = build_args()

    # ── Credentials ──────────────────────────────────────────────────────────
    password = args.password or getpass("FMC Password: ")
    console.print()

    # ── Authenticate ─────────────────────────────────────────────────────────
    with console.status("[bold cyan]Authenticating to FMC…[/bold cyan]", spinner="dots"):
        session = authenticate(args.server, args.username, password, args.cert_path)
    console.print("[green]✓ Authentication successful.[/green]")
    console.print()

    token    = session["token"]
    cert     = args.cert_path

    # ── Discover domains ─────────────────────────────────────────────────────
    with console.status("[bold cyan]Discovering domains…[/bold cyan]", spinner="dots"):
        raw_domains = get_all_domains(args.server, token, cert)

    domains = build_domain_tree(raw_domains)
    console.print(f"[green]✓ Found [bold]{len(domains)}[/bold] domain(s).[/green]")
    console.print()

    if not domains:
        console.print("[red]No domains returned. Exiting.[/red]")
        sys.exit(1)

    # ── Domain selection ─────────────────────────────────────────────────────
    console.print(Rule("[bold cyan]Domain Selection[/bold cyan]"))
    console.print()
    chosen_domains = prompt_domain_selection(domains)

    console.print()
    console.print(f"[bold green]You selected {len(chosen_domains)} domain(s):[/bold green]")
    for d in chosen_domains:
        console.print(f"  • [cyan]{d['name']}[/cyan]")
    console.print()

    if not Confirm.ask("[bold]Proceed with parsing?[/bold]", default=True):
        console.print("[yellow]Aborted.[/yellow]")
        sys.exit(0)

    console.print()
    console.print(Rule("[bold cyan]Fetching Objects[/bold cyan]"))
    console.print()

    # ── Fetch objects ─────────────────────────────────────────────────────────
    all_rows, unused_rows, summary = parse_domains(args.server, token, chosen_domains, cert)

    # ── Write CSV ─────────────────────────────────────────────────────────────
    console.print()
    console.print(Rule("[bold cyan]Exporting[/bold cyan]"))
    console.print()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    domains_queried = [d["name"] for d in chosen_domains]

    if args.output:
        out_path = args.output if os.path.isabs(args.output) else os.path.join(OUTPUT_DIR, args.output)
        if not out_path.endswith(".xlsx"):
            out_path += ".xlsx"
    else:
        safe_name = "multi_domain" if len(chosen_domains) > 1 else chosen_domains[0]["name"].replace("/", "_").replace(" ", "_")
        out_path = os.path.join(OUTPUT_DIR, f"FMC_Objects_{safe_name}_{ts}.xlsx")

    with console.status("[bold cyan]Writing Excel workbook…[/bold cyan]", spinner="dots"):
        write_excel(all_rows, unused_rows, summary, out_path, domains_queried, generated_at)

    console.print(f"[green]✓ Excel workbook written:[/green] [bold white]{out_path}[/bold white]")
    console.print(f"[green]✓ Total rows: [bold yellow]{len(all_rows)}[/bold yellow][/green]")
    console.print()

    # ── Summary table ─────────────────────────────────────────────────────────
    display_summary_table(summary)
    console.print()

    # ── Logout ────────────────────────────────────────────────────────────────
    if not args.no_logout:
        with console.status("[bold cyan]Closing FMC session…[/bold cyan]", spinner="dots"):
            logout(args.server, token, cert)
        console.print("[green]✓ FMC session closed.[/green]")

    console.print()
    console.print(Panel(
        f"[bold green]Done![/bold green]  Excel workbook saved to:\n[dim]{out_path}[/dim]",
        border_style="green",
        padding=(0, 2),
    ))


if __name__ == "__main__":
    main()
